import openpyxl
import json
from openpyxl import Workbook
from os.path import join

def excel_to_json(excel: str, json_path: str= '', attr_row: int= 2):
    """
    convert a excel file(only support xlsx, xlsm, xltx, xltm) into a JSON file
    every single worksheet will be convert into a single JSON file
    every single worksheet should have a attribute row

    for example:
        excel:
            attr_a  attr_b
            1       s
            2       ss
        will be JSON:
            [
                {
                    "attr_a": 1,
                    "attr_b": "s"
                },
                {
                    "attr_a": 2,
                    "attr_b": "ss"
                }
            ]
        attribute list is ['attr_a', 'attr_b'] and row below will be attribute's value
    
    return a dict which it's key is worksheet's title and it's value is dict (attributes:data_type) 
    """
    wb = openpyxl.load_workbook(excel)
    a_dic = {} 
    for ws in wb:
        if (ws.max_row <= attr_row):
            continue
        attrs = []
        col = 1
        while col <= ws.max_column:
            c = ws.cell(row=attr_row, column=col) 
            attrs.append(c.value)
            col += 1
        print("attrs: " + str(attrs))
       
        entities = []
        a_t = {}
        row, col = attr_row+1, 1
        while row <= ws.max_row:
            d = {}
            while col <= ws.max_column:
                c = ws.cell(row=row, column=col)
                a = str(attrs[col - 1]) 
                d[a] = c.value
                col += 1
                if (c.value != None and a not in a_t):
                    a_t[a] = type(c.value)
            entities.append(d)
            row, col = row+1, 1
        a_dic[ws.title] = a_t 

        jp = join(json_path, ws.title + ".json")
        with open(jp, 'w') as f:
            json.dump(entities, f, indent=4)
            print("dump to " + jp) 
    return a_dic

def json_to_excel(jsons:list[str], excel:str):
    wb = Workbook()
    for j in jsons:
        with open(j) as f:
            j_obj = json.load(f)
        print(f"parse {j}...")
        title = j[:j.index('.')]
        if (isinstance(j_obj, list)):
            parse_list(wb, title, j_obj)
        elif (isinstance(j_obj, dict)):
            parse_dict(wb, title, j_obj)
    if (len(wb.sheetnames) > 1):
        del wb[wb.sheetnames[0]]
    wb.save(excel)

def parse_dict(wb, title, d):
    wb.create_sheet(title=title)
    ws = wb[title]
    index_map = {}
    print(f"prase dict: {title}...")
    parse_name_values(wb, ws, index_map, 2, d)
    fill_attributes(ws, index_map)

def fill_attributes(ws, index_map):
    for k,v in index_map.items():
        ws.cell(row=1, column=v, value=k)

def parse_list(wb, title, l):
    wb.create_sheet(title=title)
    ws = wb[title]
    index_map = {}
    print(f"prase list: {title}...")
    cur_row = 2
    for o in l:
        if (isinstance(o, dict) == False):
            continue
        parse_name_values(wb, ws, index_map, cur_row, o)
        cur_row += 1   
    fill_attributes(ws=ws, index_map=index_map)

def parse_name_values(wb, ws, index_map, cur_row, dict_obj):
    for k,v in dict_obj.items():
        if (isinstance(v, list)):
            parse_list(wb, k, v)
        elif (isinstance(v, dict)):
            parse_dict(wb, k, v)
        else:
            if (k not in index_map):
                index_map[k] = len(index_map) + 1
            cur_col = index_map[k]
            ws.cell(row=cur_row,column=cur_col, value=v)